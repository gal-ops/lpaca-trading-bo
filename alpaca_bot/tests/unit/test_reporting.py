"""Unit tests for the 13-sheet Excel report and the monitoring CLI
(spec sections 13-14)."""

from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook

from alpaca_bot.persistence.db import Database
from alpaca_bot.reporting.excel_report import SHEET_NAMES, generate_report
from alpaca_bot.reporting.monitor_cli import render_status
from alpaca_bot.universe.discovery import AssetRecord


@pytest.fixture
def db(tmp_path):
    database = Database(str(tmp_path / "test.db"))
    yield database
    database.close()


def _seed(db: Database) -> None:
    db.upsert_assets([AssetRecord(
        asset_id="id-AAPL", symbol="AAPL", name="Apple", asset_class="us_equity",
        exchange="NASDAQ", status="active", tradable=True, fractionable=True,
        marginable=True, shortable=True, easy_to_borrow=True,
        maintenance_margin_requirement=0.25, min_order_increment=1,
        min_trade_increment=1, price_increment=0.01, last_checked=datetime.now(timezone.utc),
    )])
    db.record_signal({
        "signal_id": "sig-1", "strategy": "vwap_pullback_continuation", "symbol": "AAPL",
        "asset_class": "stock", "direction": "long", "regime": "BULL_TREND",
        "entry": 100, "stop": 98, "target": 104, "calibrated_probability": 0.9,
        "expected_value_after_costs": 1.5, "accepted": True,
    })
    db.record_signal_outcome("sig-1", True)
    db.record_signal({
        "signal_id": "sig-2", "strategy": "vwap_pullback_continuation", "symbol": "MSFT",
        "asset_class": "stock", "direction": "long", "regime": "BULL_TREND",
        "accepted": False, "rejection_reasons": ["reward/risk too low"],
    })
    db.record_order({
        "client_order_id": "coid-1", "broker_order_id": "b-1", "symbol": "AAPL",
        "asset_class": "stock", "side": "buy", "order_type": "market", "qty": 1,
        "status": "filled", "signal_id": "sig-1",
    })
    db.record_fill("coid-1", "AAPL", "buy", 1, 100.5)
    db.upsert_position({"symbol": "AAPL", "asset_class": "stock", "qty": 1, "avg_entry_price": 100.5})
    db.record_pnl_snapshot({"equity": 540.0, "cash": 440.0, "open_positions": 1, "gross_exposure_pct": 0.18})
    db.record_pnl_snapshot({"equity": 545.0, "cash": 445.0, "open_positions": 1, "gross_exposure_pct": 0.18})
    db.record_regime("equity", "BULL_TREND", {"breadth_pct": 65})
    db.upsert_calibration_bucket("vwap_pullback_continuation|long|stock|BULL_TREND", n_examples=400)
    db.record_risk_event("daily_stop", False, ["daily loss limit reached"])
    db.record_error("test_component", "something went wrong")


def test_generate_report_creates_all_13_sheets(db, tmp_path):
    _seed(db)
    output_path = str(tmp_path / "report.xlsx")
    generate_report(db, output_path)

    wb = load_workbook(output_path)
    assert set(wb.sheetnames) == set(SHEET_NAMES)
    assert len(SHEET_NAMES) == 13
    assert wb.active.title == "Dashboard"


def test_generate_report_trades_sheet_has_fill_data(db, tmp_path):
    _seed(db)
    output_path = str(tmp_path / "report.xlsx")
    generate_report(db, output_path)
    wb = load_workbook(output_path)
    ws = wb["Trades"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "Timestamp"
    assert any(r[1] == "AAPL" for r in rows[1:])


def test_generate_report_signals_split_accepted_and_rejected(db, tmp_path):
    _seed(db)
    output_path = str(tmp_path / "report.xlsx")
    generate_report(db, output_path)
    wb = load_workbook(output_path)
    accepted_rows = list(wb["Signals_Accepted"].iter_rows(values_only=True))
    rejected_rows = list(wb["Signals_Rejected"].iter_rows(values_only=True))
    assert any(r[2] == "AAPL" for r in accepted_rows[1:])
    assert any(r[2] == "MSFT" for r in rejected_rows[1:])


def test_generate_report_on_empty_database_does_not_crash(db, tmp_path):
    output_path = str(tmp_path / "empty_report.xlsx")
    generate_report(db, output_path)  # no seed data at all
    wb = load_workbook(output_path)
    assert set(wb.sheetnames) == set(SHEET_NAMES)


def test_generate_report_confidence_calibration_sheet(db, tmp_path):
    _seed(db)
    output_path = str(tmp_path / "report.xlsx")
    generate_report(db, output_path)
    wb = load_workbook(output_path)
    rows = list(wb["Confidence_Calibration"].iter_rows(values_only=True))
    assert rows[0][0] == "Bucket"
    assert any("vwap_pullback_continuation" in str(r[0]) for r in rows[1:])


def test_generate_report_risk_events_and_errors_sheets(db, tmp_path):
    _seed(db)
    output_path = str(tmp_path / "report.xlsx")
    generate_report(db, output_path)
    wb = load_workbook(output_path)
    risk_rows = list(wb["Risk_Events"].iter_rows(values_only=True))
    error_rows = list(wb["Errors"].iter_rows(values_only=True))
    assert any(r[1] == "daily_stop" for r in risk_rows[1:])
    assert any(r[1] == "test_component" for r in error_rows[1:])


# ---- monitoring CLI ----

def test_render_status_includes_core_fields(db):
    _seed(db)
    snapshot = {
        "account_status": "ACTIVE", "feed_type": "IEX", "market_data_healthy": True,
        "trading_stream_healthy": True, "equity": 545.0, "cash": 445.0, "gross_exposure_pct": 0.18,
    }
    output = render_status(db, snapshot)
    assert "ACTIVE" in output
    assert "$545.00" in output
    assert "BULL_TREND" in output
    assert "AAPL" in output


def test_render_status_refuses_snapshot_with_credential_keys(db):
    with pytest.raises(ValueError, match="credential"):
        render_status(db, {"api_key": "PK123", "equity": 500})


def test_render_status_works_on_empty_database(db):
    output = render_status(db, {"equity": 0, "cash": 0})
    assert "unknown" in output  # regimes not yet classified
    assert "(none)" in output  # no accepted candidates yet
