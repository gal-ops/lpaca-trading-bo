"""The 13-sheet Excel report (spec section 13). SQLite is the source of
truth; this module only reads from it and renders a workbook -- it never
writes decisions back into the database.

Generated at the end of each equity session and at 00:05 UTC for crypto
per the spec's schedule; the actual scheduling trigger lives in main.py /
a scheduled script (scripts/export_excel.py), not here -- this module is
just "build the workbook from whatever's in the DB right now."
"""

from __future__ import annotations

import json
import statistics

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font

from alpaca_bot.persistence.db import Database
from alpaca_bot.reporting.style import autosize, pnl_conditional_format, style_header, write_table

SHEET_NAMES = [
    "Dashboard", "Trades", "Open_Positions", "Signals_Accepted", "Signals_Rejected",
    "Daily_PnL", "Equity_Curve", "Asset_Universe", "Market_Regimes", "Strategy_Stats",
    "Confidence_Calibration", "Risk_Events", "Errors",
]


def generate_report(db: Database, output_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    equity_curve = _build_equity_curve(wb, db)
    trades_stats = _build_trades(wb, db)
    _build_open_positions(wb, db)
    _build_signals(wb, db, accepted=True)
    _build_signals(wb, db, accepted=False)
    _build_daily_pnl(wb, db)
    _build_asset_universe(wb, db)
    _build_market_regimes(wb, db)
    _build_strategy_stats(wb, db)
    _build_confidence_calibration(wb, db)
    _build_risk_events(wb, db)
    _build_errors(wb, db)
    _build_dashboard(wb, db, equity_curve, trades_stats)  # last: reads from the sheets above

    wb.active = wb.sheetnames.index("Dashboard")
    wb.save(output_path)


def _build_dashboard(wb, db: Database, equity_curve: list[float], trades_stats: dict) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    latest_pnl = db.query_one("SELECT * FROM pnl_snapshots ORDER BY ts DESC LIMIT 1")
    equity_regime = db.latest_regime("equity")
    crypto_regime = db.latest_regime("crypto")

    max_dd = _max_drawdown_pct(equity_curve) if equity_curve else 0.0
    sharpe = _sharpe(_period_returns(equity_curve)) if len(equity_curve) > 1 else 0.0

    rows = [
        ["Equity ($)", latest_pnl["equity"] if latest_pnl else None],
        ["Cash ($)", latest_pnl["cash"] if latest_pnl else None],
        ["Gross Exposure (%)", (latest_pnl["gross_exposure_pct"] if latest_pnl else None)],
        ["Open Positions", latest_pnl["open_positions"] if latest_pnl else None],
        ["Equity Regime", equity_regime["regime"] if equity_regime else "unknown"],
        ["Crypto Regime", crypto_regime["regime"] if crypto_regime else "unknown"],
        ["Total Trades", trades_stats["n_trades"]],
        ["Win Rate (%)", trades_stats["win_rate"] * 100],
        ["Total Realized P&L ($)", trades_stats["total_pnl"]],
        ["Max Drawdown (%)", max_dd],
        ["Sharpe (approx.)", sharpe],
    ]
    ws.append(["Metric", "Value"])
    for row_values in rows:
        ws.append(row_values)
    style_header(ws, ncols=2)
    autosize(ws, [26, 20])
    for row_num in range(2, ws.max_row + 1):
        ws.cell(row=row_num, column=1).font = Font(bold=True)


def _build_trades(wb, db: Database) -> dict:
    ws = wb.create_sheet("Trades")
    fills = db.query("""
        SELECT f.ts, f.symbol, f.side, f.qty, f.price, o.order_type, o.client_order_id,
               s.strategy, s.regime
        FROM fills f
        LEFT JOIN orders o ON o.client_order_id = f.client_order_id
        LEFT JOIN signals s ON s.signal_id = o.signal_id
        ORDER BY f.ts
    """)
    headers = ["Timestamp", "Symbol", "Side", "Qty", "Price", "Order Type",
               "Client Order ID", "Strategy", "Regime"]
    rows = [[r["ts"], r["symbol"], r["side"], r["qty"], r["price"], r["order_type"],
             r["client_order_id"], r["strategy"], r["regime"]] for r in fills]
    write_table(ws, headers, rows, widths=[20, 10, 8, 10, 12, 12, 24, 24, 18])

    outcomes = db.query("SELECT outcome_label FROM signals WHERE outcome_label IS NOT NULL")
    n = len(outcomes)
    wins = sum(1 for o in outcomes if o["outcome_label"] == 1)
    return {
        "n_trades": len(fills), "win_rate": (wins / n) if n else 0.0,
        "total_pnl": None,  # true realized P&L needs FIFO trade-matching, not yet wired in phase 10
    }


def _build_open_positions(wb, db: Database) -> None:
    ws = wb.create_sheet("Open_Positions")
    rows = db.open_positions()
    headers = ["Symbol", "Asset Class", "Qty", "Avg Entry", "Current Price", "Market Value",
               "Unrealized P&L ($)", "Unrealized P&L (%)", "Strategy", "Stop", "Target", "Opened At"]
    data = [[r["symbol"], r["asset_class"], r["qty"], r["avg_entry_price"], r["current_price"],
             r["market_value"], r["unrealized_pl"], r["unrealized_plpc"], r["strategy"],
             r["stop"], r["target"], r["opened_at"]] for r in rows]
    write_table(ws, headers, data, widths=[10, 11, 8, 11, 13, 13, 16, 16, 20, 9, 9, 20])
    pnl_conditional_format(ws, "G", max_row=500)


def _build_signals(wb, db: Database, accepted: bool) -> None:
    name = "Signals_Accepted" if accepted else "Signals_Rejected"
    ws = wb.create_sheet(name)
    rows = db.query(
        "SELECT * FROM signals WHERE accepted = ? ORDER BY ts DESC LIMIT 5000", (int(accepted),)
    )
    headers = ["Timestamp", "Strategy", "Symbol", "Asset Class", "Direction", "Regime",
               "Entry", "Stop", "Target", "Calibrated Probability", "Expected Value After Costs"]
    if not accepted:
        headers.append("Rejection Reasons")
    data = []
    for r in rows:
        row = [r["ts"], r["strategy"], r["symbol"], r["asset_class"], r["direction"], r["regime"],
               r["entry"], r["stop"], r["target"], r["calibrated_probability"],
               r["expected_value_after_costs"]]
        if not accepted:
            reasons = json.loads(r["rejection_reasons_json"]) if r["rejection_reasons_json"] else []
            row.append(", ".join(reasons))
        data.append(row)
    widths = [20, 24, 10, 11, 10, 18, 10, 10, 10, 14, 14] + ([40] if not accepted else [])
    write_table(ws, headers, data, widths=widths)


def _build_daily_pnl(wb, db: Database) -> None:
    ws = wb.create_sheet("Daily_PnL")
    snapshots = db.query("SELECT ts, equity FROM pnl_snapshots ORDER BY ts")
    by_day: dict[str, list[float]] = {}
    for row in snapshots:
        day = row["ts"][:10]
        by_day.setdefault(day, []).append(row["equity"])
    rows = []
    prev_close = None
    for day in sorted(by_day):
        values = by_day[day]
        open_eq, close_eq = values[0], values[-1]
        change = (close_eq - prev_close) if prev_close is not None else 0.0
        change_pct = (change / prev_close * 100) if prev_close else 0.0
        rows.append([day, open_eq, close_eq, round(change, 2), round(change_pct, 4)])
        prev_close = close_eq
    write_table(ws, ["Date", "Open Equity", "Close Equity", "Change ($)", "Change (%)"],
                rows, widths=[14, 14, 14, 12, 12])
    pnl_conditional_format(ws, "D", max_row=len(rows) + 1)


def _build_equity_curve(wb, db: Database) -> list[float]:
    ws = wb.create_sheet("Equity_Curve")
    rows = db.query("SELECT ts, equity, cash FROM pnl_snapshots ORDER BY ts")
    write_table(ws, ["Timestamp", "Equity", "Cash"],
                [[r["ts"], r["equity"], r["cash"]] for r in rows], widths=[20, 14, 14])
    if len(rows) > 1:
        chart = LineChart()
        chart.title = "Equity Curve"
        chart.y_axis.title = "Equity ($)"
        chart.x_axis.title = "Cycle #"
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.width, chart.height = 24, 10
        ws.add_chart(chart, "E2")
    return [r["equity"] for r in rows]


def _build_asset_universe(wb, db: Database) -> None:
    ws = wb.create_sheet("Asset_Universe")
    rows = db.query("""
        SELECT symbol, asset_class, exchange, status, tradable, fractionable, marginable,
               shortable, easy_to_borrow, last_checked, exclusion_reason
        FROM assets ORDER BY asset_class, symbol LIMIT 20000
    """)
    headers = ["Symbol", "Asset Class", "Exchange", "Status", "Tradable", "Fractionable",
               "Marginable", "Shortable", "Easy To Borrow", "Last Checked", "Exclusion Reason"]
    data = [[r["symbol"], r["asset_class"], r["exchange"], r["status"], bool(r["tradable"]),
             bool(r["fractionable"]), bool(r["marginable"]), bool(r["shortable"]),
             bool(r["easy_to_borrow"]), r["last_checked"], r["exclusion_reason"]] for r in rows]
    write_table(ws, headers, data, widths=[10, 11, 10, 9, 9, 11, 10, 9, 13, 20, 30])


def _build_market_regimes(wb, db: Database) -> None:
    ws = wb.create_sheet("Market_Regimes")
    rows = db.query("SELECT asset_class, ts, regime, inputs_json FROM regimes ORDER BY ts DESC LIMIT 5000")
    write_table(ws, ["Asset Class", "Timestamp", "Regime", "Inputs"],
                [[r["asset_class"], r["ts"], r["regime"], r["inputs_json"]] for r in rows],
                widths=[11, 20, 24, 60])


def _build_strategy_stats(wb, db: Database) -> None:
    ws = wb.create_sheet("Strategy_Stats")
    rows = db.query("""
        SELECT strategy, direction, asset_class, regime,
               SUM(accepted) AS n_accepted,
               SUM(CASE WHEN accepted = 0 THEN 1 ELSE 0 END) AS n_rejected,
               SUM(CASE WHEN outcome_label = 1 THEN 1 ELSE 0 END) AS n_wins,
               SUM(CASE WHEN outcome_label = 0 THEN 1 ELSE 0 END) AS n_losses,
               AVG(calibrated_probability) AS avg_calibrated_probability,
               AVG(expected_value_after_costs) AS avg_expected_value
        FROM signals
        GROUP BY strategy, direction, asset_class, regime
        ORDER BY strategy, direction, asset_class, regime
    """)
    headers = ["Strategy", "Direction", "Asset Class", "Regime", "Accepted", "Rejected",
               "Wins", "Losses", "Win Rate (%)", "Avg Calibrated Probability", "Avg Expected Value"]
    data = []
    for r in rows:
        n_decided = (r["n_wins"] or 0) + (r["n_losses"] or 0)
        win_rate = (r["n_wins"] / n_decided * 100) if n_decided else None
        data.append([
            r["strategy"], r["direction"], r["asset_class"], r["regime"], r["n_accepted"],
            r["n_rejected"], r["n_wins"], r["n_losses"], win_rate,
            r["avg_calibrated_probability"], r["avg_expected_value"],
        ])
    write_table(ws, headers, data, widths=[24, 10, 11, 18, 10, 10, 8, 8, 12, 20, 16])


def _build_confidence_calibration(wb, db: Database) -> None:
    ws = wb.create_sheet("Confidence_Calibration")
    rows = db.query("SELECT * FROM calibration_buckets ORDER BY bucket_key")
    headers = ["Bucket", "N Examples", "Model Version", "Disabled", "Disabled Reason",
               "Observed Hit Rate Near 85% (%)", "Sample Size Near 85%"]
    data = []
    for r in rows:
        near_85 = db.bucket_outcomes_near_probability(r["bucket_key"], 0.85, 0.05)
        n_near = len(near_85)
        hit_rate = (sum(o["outcome_label"] for o in near_85) / n_near * 100) if n_near else None
        data.append([r["bucket_key"], r["n_examples"], r["model_version"], bool(r["disabled"]),
                     r["disabled_reason"], hit_rate, n_near])
    write_table(ws, headers, data, widths=[40, 12, 20, 10, 40, 24, 18])


def _build_risk_events(wb, db: Database) -> None:
    ws = wb.create_sheet("Risk_Events")
    rows = db.query("SELECT * FROM risk_events ORDER BY ts DESC LIMIT 5000")
    data = []
    for r in rows:
        reasons = json.loads(r["reasons_json"]) if r["reasons_json"] else []
        data.append([r["ts"], r["event_type"], bool(r["should_flatten"]), "; ".join(reasons)])
    write_table(ws, ["Timestamp", "Event Type", "Should Flatten", "Reasons"], data,
                widths=[20, 20, 14, 60])


def _build_errors(wb, db: Database) -> None:
    ws = wb.create_sheet("Errors")
    rows = db.query("SELECT * FROM errors ORDER BY ts DESC LIMIT 5000")
    write_table(ws, ["Timestamp", "Component", "Message", "Traceback"],
                [[r["ts"], r["component"], r["message"], r["traceback"]] for r in rows],
                widths=[20, 20, 50, 60])


# ---- shared small-sample metric helpers (mirrors backtester.metrics, kept
# local/lightweight here to avoid this reporting module depending on the
# backtester package) ----

def _max_drawdown_pct(equity_values: list[float]) -> float:
    if not equity_values:
        return 0.0
    peak = equity_values[0]
    max_dd = 0.0
    for v in equity_values:
        peak = max(peak, v)
        if peak > 0:
            max_dd = max(max_dd, (peak - v) / peak)
    return max_dd * 100


def _period_returns(equity_values: list[float]) -> list[float]:
    returns = []
    for i in range(1, len(equity_values)):
        prev = equity_values[i - 1]
        if prev:
            returns.append((equity_values[i] - prev) / prev)
    return returns


def _sharpe(returns: list[float]) -> float:
    if len(returns) < 2:
        return 0.0
    mean = statistics.mean(returns)
    stdev = statistics.pstdev(returns)
    return (mean / stdev) if stdev else 0.0
