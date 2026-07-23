"""Shared openpyxl styling helpers so all 13 report sheets look like one
consistent workbook instead of 13 independently-styled ones."""

from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
POS_FILL = PatternFill("solid", fgColor="C6EFCE")
POS_FONT = Font(color="006100")
NEG_FILL = PatternFill("solid", fgColor="FFC7CE")
NEG_FONT = Font(color="9C0006")


def write_table(ws, headers: list[str], rows: list[list], widths: list[int] | None = None) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws, ncols=len(headers))
    if widths:
        autosize(ws, widths)


def style_header(ws, row: int = 1, ncols: int | None = None) -> None:
    ncols = ncols or ws.max_column
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"A{row + 1}"


def autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def pnl_conditional_format(ws, col_letter: str, max_row: int = 100000) -> None:
    if max_row < 2:
        return  # no data rows to format
    rng = f"{col_letter}2:{col_letter}{max_row}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="greaterThan", formula=["0"], fill=POS_FILL, font=POS_FONT))
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="lessThan", formula=["0"], fill=NEG_FILL, font=NEG_FONT))
