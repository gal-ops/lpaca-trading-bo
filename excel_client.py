import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, BarChart, PieChart, DoughnutChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import config

TRADES_HEADERS = [
    "Timestamp", "Symbol", "Asset Class", "Side", "Qty", "Price", "Value ($)",
    "Reason", "Realized P&L ($)", "Realized P&L (%)", "Order ID",
    "Cumulative P&L ($)",
]

POSITIONS_HEADERS = [
    "Symbol", "Asset Class", "Qty", "Avg Entry Price", "Current Price",
    "Market Value ($)", "Cost Basis ($)", "Unrealized P&L ($)", "Unrealized P&L (%)",
]

EQUITY_LOG_HEADERS = [
    "Timestamp", "Equity ($)", "Cash ($)", "Buying Power ($)",
    "Daily P&L ($)", "Daily P&L (%)", "High-Water Mark ($)", "Drawdown (%)",
    "Open Positions", "Exposure ($)", "Exposure (%)",
]

# Fixed row ranges used for chart/formula references so they keep working
# as more rows get appended, without having to rebuild charts each run.
MAX_ROWS = 100000
SYMBOL_TABLE_MAX_ROWS = 200
SYMBOL_TABLE_START_ROW = 2

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
POS_FILL = PatternFill("solid", fgColor="C6EFCE")
POS_FONT = Font(color="006100")
NEG_FILL = PatternFill("solid", fgColor="FFC7CE")
NEG_FONT = Font(color="9C0006")


def _style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"A{row + 1}"


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _pnl_conditional_format(ws, col_letter, max_row=MAX_ROWS):
    rng = f"{col_letter}2:{col_letter}{max_row}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="greaterThan", formula=["0"], fill=POS_FILL, font=POS_FONT))
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="lessThan", formula=["0"], fill=NEG_FILL, font=NEG_FONT))


def _ensure_trades_sheet(wb):
    is_new = "Trades" not in wb.sheetnames
    ws = wb.create_sheet("Trades") if is_new else wb["Trades"]
    if is_new:
        ws.append(TRADES_HEADERS)
        _style_header(ws, ncols=len(TRADES_HEADERS))
        _autosize(ws, [19, 10, 11, 6, 10, 11, 12, 34, 16, 16, 20, 18])
        _pnl_conditional_format(ws, "I")
        _pnl_conditional_format(ws, "L")
    elif [c.value for c in ws[1]] != TRADES_HEADERS:
        for col, header in enumerate(TRADES_HEADERS, start=1):
            ws.cell(row=1, column=col, value=header)
    return ws


def _ensure_positions_sheet(wb):
    is_new = "Positions" not in wb.sheetnames
    ws = wb.create_sheet("Positions") if is_new else wb["Positions"]
    if is_new:
        ws.append(POSITIONS_HEADERS)
        _style_header(ws, ncols=len(POSITIONS_HEADERS))
        _autosize(ws, [10, 11, 10, 15, 14, 14, 13, 16, 16])
        _pnl_conditional_format(ws, "H", max_row=500)
    return ws


def _ensure_equity_log_sheet(wb):
    is_new = "EquityLog" not in wb.sheetnames
    ws = wb.create_sheet("EquityLog") if is_new else wb["EquityLog"]
    if is_new:
        ws.append(EQUITY_LOG_HEADERS)
        _style_header(ws, ncols=len(EQUITY_LOG_HEADERS))
        _autosize(ws, [19, 13, 12, 14, 13, 12, 17, 12, 14, 12, 12])
        _pnl_conditional_format(ws, "E")
        _pnl_conditional_format(ws, "H")
    return ws


def _ensure_summary_sheet(wb):
    if "Summary" in wb.sheetnames:
        return wb["Summary"]
    ws = wb.create_sheet("Summary")
    ws.append(["Metric", "Value"])
    ws.append(["Total Trades", f"=COUNTA(Trades!A2:A{MAX_ROWS})"])
    ws.append(["Wins", f'=COUNTIF(Trades!I2:I{MAX_ROWS},">0")'])
    ws.append(["Losses", f'=COUNTIF(Trades!I2:I{MAX_ROWS},"<0")'])
    ws.append(["Win Rate (%)", "=IFERROR(B3/(B3+B4)*100,0)"])
    ws.append(["Total Realized P&L ($)", f"=SUM(Trades!I2:I{MAX_ROWS})"])
    _style_header(ws, ncols=2)
    _autosize(ws, [24, 20])
    return ws


def _ensure_dashboard_sheet(wb):
    is_new = "Dashboard" not in wb.sheetnames
    ws = wb["Dashboard"] if not is_new else wb.create_sheet("Dashboard")
    if not is_new:
        return ws

    # Pulls the latest row of EquityLog by position (COUNTA of the timestamp
    # column), so these stay live without a fixed row reference as the log
    # grows every cycle.
    last = f"COUNTA(EquityLog!A:A)"
    def latest(col):
        return f"=INDEX(EquityLog!{col}:{col},{last})"

    rows = [
        ("Last Updated", latest("A")),
        ("Equity ($)", latest("B")),
        ("Cash ($)", latest("C")),
        ("Buying Power ($)", latest("D")),
        ("Daily P&L ($)", latest("E")),
        ("Daily P&L (%)", latest("F")),
        ("High-Water Mark ($)", latest("G")),
        ("Drawdown (%)", latest("H")),
        ("Open Positions", latest("I")),
        ("Exposure ($)", latest("J")),
        ("Exposure (%)", latest("K")),
        ("", ""),
        ("Total Trades", "=Summary!B2"),
        ("Win Rate (%)", "=Summary!B5"),
        ("Total Realized P&L ($)", "=Summary!B6"),
    ]
    ws.append(["Metric", "Value"])
    for label, formula in rows:
        ws.append([label, formula])
    _style_header(ws, ncols=2)
    _autosize(ws, [26, 20])
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)
    return ws


def _ensure_charts_sheet(wb):
    is_new = "Charts" not in wb.sheetnames
    ws = wb["Charts"] if not is_new else wb.create_sheet("Charts")

    if is_new:
        ws["A1"] = "Symbol"
        ws["B1"] = "Total Realized P&L ($)"
        ws["D1"] = "Asset Class"
        ws["E1"] = "Trade Count"

        # --- Line chart: cumulative P&L trend over every trade ---
        line = LineChart()
        line.title = "Cumulative P&L Over Time"
        line.y_axis.title = "P&L ($)"
        line.x_axis.title = "Trade #"
        line.style = 12
        data = Reference(wb["Trades"], min_col=12, min_row=1, max_row=MAX_ROWS)
        line.add_data(data, titles_from_data=True)
        line.width, line.height = 24, 10
        ws.add_chart(line, "H2")

        # --- Line chart: account equity curve over every logged cycle ---
        equity = LineChart()
        equity.title = "Account Equity Over Time"
        equity.y_axis.title = "Equity ($)"
        equity.x_axis.title = "Cycle #"
        equity.style = 13
        data = Reference(wb["EquityLog"], min_col=2, min_row=1, max_row=MAX_ROWS)
        equity.add_data(data, titles_from_data=True)
        equity.width, equity.height = 24, 10
        ws.add_chart(equity, "H22")

        # --- Pie chart: win vs loss distribution ---
        pie = PieChart()
        pie.title = "Win / Loss Distribution"
        cats = Reference(wb["Summary"], min_col=1, min_row=3, max_row=4)
        data = Reference(wb["Summary"], min_col=2, min_row=3, max_row=4)
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(cats)
        pie.width, pie.height = 12, 10
        ws.add_chart(pie, "A22")

        # --- Bar chart: realized P&L by symbol ---
        bar = BarChart()
        bar.type = "col"
        bar.title = "Realized P&L by Symbol"
        bar.y_axis.title = "P&L ($)"
        bar.x_axis.title = "Symbol"
        data = Reference(ws, min_col=2, min_row=1, max_row=SYMBOL_TABLE_MAX_ROWS + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=SYMBOL_TABLE_MAX_ROWS + 1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.width, bar.height = 24, 10
        ws.add_chart(bar, "A42")

        # --- Doughnut chart: trades by asset class (stock vs crypto) ---
        donut = DoughnutChart()
        donut.title = "Trades by Asset Class"
        data = Reference(ws, min_col=5, min_row=1, max_row=3)
        cats = Reference(ws, min_col=4, min_row=2, max_row=3)
        donut.add_data(data, titles_from_data=True)
        donut.set_categories(cats)
        donut.width, donut.height = 12, 10
        ws.add_chart(donut, "A60")

    return ws


def _refresh_chart_tables(wb):
    """Recompute the small aggregation tables the bar/doughnut charts read from."""
    trades_ws = wb["Trades"]
    charts_ws = wb["Charts"]

    symbol_pnl = {}
    asset_class_counts = {}
    for row in trades_ws.iter_rows(min_row=2, values_only=True):
        symbol, asset_class, pnl = row[1], row[2], row[8]
        if asset_class:
            asset_class_counts[asset_class] = asset_class_counts.get(asset_class, 0) + 1
        if pnl not in (None, ""):
            symbol_pnl[symbol] = symbol_pnl.get(symbol, 0) + float(pnl)

    for r in range(SYMBOL_TABLE_START_ROW, SYMBOL_TABLE_START_ROW + SYMBOL_TABLE_MAX_ROWS):
        charts_ws.cell(row=r, column=1, value=None)
        charts_ws.cell(row=r, column=2, value=None)

    for i, (symbol, total) in enumerate(sorted(symbol_pnl.items(), key=lambda kv: -kv[1])):
        if i >= SYMBOL_TABLE_MAX_ROWS:
            break
        r = SYMBOL_TABLE_START_ROW + i
        charts_ws.cell(row=r, column=1, value=symbol)
        charts_ws.cell(row=r, column=2, value=round(total, 2))

    for i, asset_class in enumerate(["stock", "crypto"]):
        charts_ws.cell(row=2 + i, column=4, value=asset_class)
        charts_ws.cell(row=2 + i, column=5, value=asset_class_counts.get(asset_class, 0))


def _ensure_workbook():
    if os.path.exists(config.EXCEL_FILE):
        wb = load_workbook(config.EXCEL_FILE)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    _ensure_trades_sheet(wb)
    _ensure_positions_sheet(wb)
    _ensure_equity_log_sheet(wb)
    _ensure_summary_sheet(wb)
    _ensure_dashboard_sheet(wb)
    _ensure_charts_sheet(wb)
    wb.active = wb.sheetnames.index("Dashboard")
    return wb


def log_trade(symbol, asset_class, side, qty, price, reason,
               realized_pnl_usd=None, realized_pnl_pct=None, order_id=""):
    wb = _ensure_workbook()
    ws = wb["Trades"]
    value = qty * price

    prev_cumulative = ws.cell(row=ws.max_row, column=12).value if ws.max_row >= 2 else 0
    cumulative = (prev_cumulative or 0) + (realized_pnl_usd or 0)

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        symbol,
        asset_class,
        side.upper(),
        qty,
        round(price, 6),
        round(value, 2),
        reason,
        round(realized_pnl_usd, 2) if realized_pnl_usd is not None else None,
        round(realized_pnl_pct, 4) if realized_pnl_pct is not None else None,
        order_id,
        round(cumulative, 2),
    ])

    _refresh_chart_tables(wb)
    wb.save(config.EXCEL_FILE)


def log_snapshot(account, positions, asset_classes=None, extra=None):
    """Log a full point-in-time account snapshot -- called once per bot
    cycle (independent of whether any trade happened), so the workbook
    reflects live state continuously rather than only on fills.

    positions: {symbol: alpaca Position} as returned by alpaca_client.get_positions().
    asset_classes: optional {symbol: "stock"|"crypto"} override; falls back
    to the same "/" or "USD" suffix heuristic used elsewhere in the bot.
    extra: optional dict with any of high_water_mark / drawdown_pct /
    exposure_usd / exposure_pct / daily_pnl_usd / daily_pnl_pct, to reuse
    values already computed this cycle instead of recomputing them here.
    """
    extra = extra or {}
    asset_classes = asset_classes or {}
    wb = _ensure_workbook()

    equity = float(account.equity)
    cash = float(account.cash)
    buying_power = float(account.buying_power)
    last_equity = float(account.last_equity)
    daily_pnl_usd = extra.get("daily_pnl_usd", equity - last_equity)
    daily_pnl_pct = extra.get("daily_pnl_pct", (daily_pnl_usd / last_equity * 100) if last_equity else 0.0)
    hwm = extra.get("high_water_mark", equity)
    drawdown_pct = extra.get("drawdown_pct", ((equity - hwm) / hwm * 100) if hwm else 0.0)
    exposure_usd = extra.get("exposure_usd", sum(
        abs(float(getattr(p, "market_value", 0) or 0)) for p in positions.values()))
    exposure_pct = extra.get("exposure_pct", (exposure_usd / equity * 100) if equity else 0.0)

    # ---- EquityLog: append one row per cycle (the time series) ----
    eq_ws = wb["EquityLog"]
    eq_ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        round(equity, 2),
        round(cash, 2),
        round(buying_power, 2),
        round(daily_pnl_usd, 2),
        round(daily_pnl_pct, 4),
        round(hwm, 2),
        round(drawdown_pct, 4),
        len(positions),
        round(exposure_usd, 2),
        round(exposure_pct, 4),
    ])

    # ---- Positions: full rewrite each cycle (current state, not history) ----
    pos_ws = wb["Positions"]
    if pos_ws.max_row > 1:
        pos_ws.delete_rows(2, pos_ws.max_row - 1)
    for symbol, pos in sorted(positions.items()):
        asset_class = asset_classes.get(symbol) or ("crypto" if ("/" in symbol or symbol.endswith("USD")) else "stock")
        qty = float(pos.qty)
        avg_entry = float(pos.avg_entry_price)
        current_price = float(getattr(pos, "current_price", 0) or 0)
        market_value = float(getattr(pos, "market_value", 0) or 0)
        cost_basis = qty * avg_entry
        unrealized_pl = float(getattr(pos, "unrealized_pl", 0) or 0)
        unrealized_plpc = float(getattr(pos, "unrealized_plpc", 0) or 0) * 100
        pos_ws.append([
            symbol, asset_class, qty, round(avg_entry, 6), round(current_price, 6),
            round(market_value, 2), round(cost_basis, 2), round(unrealized_pl, 2),
            round(unrealized_plpc, 4),
        ])

    _refresh_chart_tables(wb)
    wb.save(config.EXCEL_FILE)
